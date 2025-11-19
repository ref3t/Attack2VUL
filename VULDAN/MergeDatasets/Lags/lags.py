import os
import json
import openpyxl
import pandas as pd


# Function to load Excel data and fill CVECreated for all matching CVEIDs
def fill_cvecreated_from_json(excel_file, json_folder):
    # Check if the folder exists
    if not os.path.exists(json_folder):
        print(f"Error: The folder '{json_folder}' does not exist.")
        return
    
    if not os.path.isdir(json_folder):
        print(f"Error: The path '{json_folder}' is not a directory.")
        return
    
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    # Read the data from Excel
    excel_data = {}
    for row in range(2, sheet.max_row + 1):  # Assuming first row is header
        technique_id = sheet.cell(row=row, column=1).value
        date_created = sheet.cell(row=row, column=2).value
        cve_id = sheet.cell(row=row, column=3).value
        
        # Initialize a list for all rows of this CVEID
        if cve_id not in excel_data:
            excel_data[cve_id] = []
        
        # Append row information (row number and CVECreated value)
        excel_data[cve_id].append({'row': row, 'technique_id': technique_id, 'date_created': date_created, 'cv_created': sheet.cell(row=row, column=4).value})
    
    # Debugging: print out the loaded data from Excel
    print("Loaded CVE IDs from Excel:", excel_data.keys())
    
    # Traverse through JSON files in all subfolders
    for root, dirs, files in os.walk(json_folder):
        # Debugging: print current directory being traversed
        print(f"Checking directory: {root}")
        
        for file in files:
            # Debugging: print the file being checked
            print(f"Checking file: {file}")
            
            if file.endswith(".json"):
                cve_id = file.replace('.json', '')  # Extract CVEID from filename
                
                # Check if CVEID exists in Excel data
                if cve_id in excel_data:
                    json_file_path = os.path.join(root, file)
                    
                    # Debugging: Print if a match is found
                    print(f"Match found for CVEID: {cve_id}")
                    
                    # Open JSON file
                    try:
                        with open(json_file_path, 'r') as f:
                            json_data = json.load(f)
                        
                        # Extract dateReserved from cveMetadata
                        if 'cveMetadata' in json_data and 'dateReserved' in json_data['cveMetadata']:
                            date_reserved = json_data['cveMetadata']['dateReserved'][:10]  # Extract only the date part
                            
                            # Update all rows corresponding to this CVEID
                            for entry in excel_data[cve_id]:
                                row_to_update = entry['row']
                                if not entry['cv_created']:  # Only update if CVECreated is empty or None
                                    sheet.cell(row=row_to_update, column=4).value = date_reserved
                                    print(f"Updated CVECreated for {cve_id} in row {row_to_update} with {date_reserved}")
                                else:
                                    print(f"CVECreated for {cve_id} in row {row_to_update} already filled with {entry['cv_created']}")
                        else:
                            print(f"No dateReserved field found in JSON file for {cve_id}")
                    
                    except json.JSONDecodeError:
                        print(f"Error reading JSON file for {cve_id}")

    # Save the updated Excel file
    wb.save(excel_file)
    print(f"Excel file {excel_file} has been updated with CVECreated dates.")

# Path to your Excel file
excel_file = './Results/Lags.xlsx'

# Path to your folder containing JSON files
json_folder = './datasets/CVEList/cves'

# Run the function
# fill_cvecreated_from_json(excel_file, json_folder)


data = pd.read_excel(excel_file, sheet_name=0)

# Display the original data
print("Original Data:")
print(data)

# Group by the "Lag (Days)" column and count the number of techniques
lag_counts = data.groupby('Lag (Days)').size().reset_index(name='Count')

# Display the counts for each Lag value
print("\nCount of Techniques by Lag (Days):")
print(lag_counts)

# Optionally, save the results to a new CSV file
lag_counts.to_csv('./Results/lag_counts.csv', index=False)