from openpyxl import load_workbook, Workbook
import re
# Load the Excel workbook
file_path = "./Results/WeekNewsResultsValidation2.xlsm"  # Replace with your Excel file path
workbook = load_workbook(file_path)
sheet = workbook.active  # Access the first sheet


# Regular expression to match CVE-YYYY-NNNN pattern
cve_pattern = r"CVE-\d{4}-\d+"

# Initialize a list to store results


finalResults = [] 
# Loop through rows and columns to check for matching values and green highlight
for row in sheet.iter_rows():
    results = []
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            # Check if the value matches the CVE pattern
            if re.match(cve_pattern, cell.value):
                # Get the fill color of the cell
                fill_color = cell.fill.fgColor
                if fill_color.type == "rgb" and ( fill_color.rgb == "FFC6EFCE" or fill_color.rgb == "FFFFEB9C" ):  # Green in RGB (hex)
                    results.append(cell.value)
    finalResults.append(", ".join(results)) 

# Save the final results to a new Excel file
output_workbook = Workbook()
output_sheet = output_workbook.active

# Write the finalResults to the new workbook, each row will have a single cell with all CVEs joined
for row_index, row_data in enumerate(finalResults, start=1):
    output_sheet.cell(row=row_index, column=1, value=row_data)  # Save each joined string in a single cell


# Save the workbook to a new file
output_file_path = "./Results/ManualExtractionResultsValidation.xlsx"  # Replace with your desired output file name
output_workbook.save(output_file_path)
print(f"Results saved to {output_file_path}")